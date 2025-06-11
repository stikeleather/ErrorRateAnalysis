import sys
import getopt
import openpyxl
import re
from codon_bias import find_match, find_mismatch
from Bio.Data import CodonTable
from generate_custom_database import fasta_to_dict

def read_file_names(file_input):
	if isinstance(file_input, list):
		return file_input
	else:
		with open(file_input, 'r') as file:
			file_names = file.read().splitlines()
		return file_names


def codon_matrix(input_file ,gene_file, protein_file, tt,):

	wb_in = openpyxl.load_workbook(input_file)
	ws0 = wb_in["Sheet1"]

	sequence_dic = fasta_to_dict(gene_file)
	protein_dic = fasta_to_dict(protein_file)

	row_num_count = 0
	for www in range(1,ws0.max_row+1):
		if ws0.cell(www,1).value is not None:
			row_num_count += 1
		else:
			break
	row_num = row_num_count+1

	aa_codons = CodonTable.unambiguous_dna_by_id[tt]
	stop_codons = aa_codons.stop_codons
	aa_codons = dict(aa_codons.forward_table.items())
	for stop in stop_codons:
		aa_codons[stop] = ''

	coding_codons = (len(aa_codons) - len(stop_codons)) + 2 # +1 for range

	#Keys are amino acids and values are codons in a list
	distance_lib = dict()
	for k, v in aa_codons.items():
		if v not in distance_lib:
			distance_lib[v] = [k]
		else:
			distance_lib[v].append(k)


	aa_sites_dic = {
		'R':0,'H':0,'K':0,'D':0,'E':0,'S':0,'T':0,'N':0,'Q':0,'C':0,
		'G':0,'P':0,'A':0,'V':0,'I':0,'L':0,'M':0,'F':0,'Y':0,'W':0
		}
	
	codon_wild_table = dict()
	for key in distance_lib:
		for codon in distance_lib[key]:
			if codon not in stop_codons:
				codon_wild_table[codon] = 0

	codon_error_table = codon_wild_table.copy()

	output = f'codon_matrix_{input_file}'
	wb = openpyxl.Workbook()
	wb.create_sheet(index=0, title='Sheet1')
	wb.save(output)
	wb = openpyxl.load_workbook(output)
	ws = wb["Sheet1"]

	current_col = 2
	for key in aa_sites_dic:
		for item in distance_lib[key]:
			ws.cell(1,current_col).value = key
			ws.cell(25,current_col).value = key
			ws.cell(2,current_col).value = item
			ws.cell(26,current_col).value = item
			current_col += 1

	row_offset = 24
	current_row = 3
	for key in aa_sites_dic:
		ws.cell(current_row,1).value = key
		ws.cell(current_row+row_offset,1).value = key
		current_row += 1

	#Intialize zeroes in matrix
	for col in range(2, coding_codons):
		for row in range(3, 23):
			ws.cell(row, col).value = 0
			ws.cell(row+row_offset, col).value = 0

	digit_pattern = re.compile(r'\d+')
	dash_dot_pattern = re.compile(r'[-.]')
	dash_pattern = re.compile(r'[-]')
	#Loops over mutants; determines source aa and dest aa; finds the intersect in the matrix and adds the PSM count to the cell
	for row in range(2,row_num):
		gene = ws0.cell(row,1).value
		if 'wild' not in ws0.cell(row,1).value:
			try:
				split_gene = re.split(dash_dot_pattern,gene)
				mutation = split_gene[4]
				src = mutation[0]
				mutation_pos = int(re.findall(digit_pattern, mutation)[0])
				mutation = mutation[-1]
				if src not in aa_sites_dic.keys() or mutation not in aa_sites_dic.keys():
					raise(f"Invalid source or destination: {key}, src: {src}, dest: {mutation}")
				gene_id = split_gene[1]
				# print(gene_id)
				nt_sequence = sequence_dic[gene_id]
				protein_seq = protein_dic[gene_id]
				aa_sequence = ws0.cell(row,2).value
				mismatches = find_mismatch(protein_seq, aa_sequence)
				# print(aa_sequence)
				# print(mismatches)
				nt_codon = ''
				nt_pos = mismatches[1] * 3
				# print(nt_pos)
				if nt_pos >= 0:
					for i in range(3):
						nt_codon += nt_sequence[nt_pos-3+i]
				psm_count = ws0.cell(row,3).value
				codon_error_table[nt_codon] += psm_count
				codon_count = 0
				for nt in range((mismatches[0]*3)-3, (mismatches[2]*3), 3):
					codon_count += 1
					codon = nt_sequence[nt:nt+3]
					#print(codon)
					if codon_count != mutation_pos:
						# print(mutation_pos)
						# print(codon)
						codon_wild_table[codon] += psm_count

			except KeyError:				
				print(f"Error in codon matrix: {gene}, check the protein and gene sequence, then the genbank file entry")
				sys.exit()

			for col in range(2,coding_codons):
				for row in range(3,23):
					#Matrix intersect
					if ws.cell(2,col).value == nt_codon and ws.cell(row,1).value == mutation:
						#Existing PSM + new PSM
						ws.cell(row,col).value = ws.cell(row,col).value + psm_count

		else:
			try:
				split_gene = re.split(dash_pattern,gene)
				gene_id = split_gene[1]
				aa_sequence = ws0.cell(row,2).value
				nt_sequence = sequence_dic[gene_id]
				protein_seq = protein_dic[gene_id]
				matches = find_match(protein_seq, aa_sequence)
				psm_count = ws0.cell(row,3).value
				for nt in range((matches[0]*3),(matches[1]*3),3):
					codon = nt_sequence[nt:nt+3]
					#print(codon)
					codon_wild_table[codon] += psm_count
			except KeyError:
				print(f"Error in codon matrix: {gene}, check the protein and gene sequence, then the genbank file entry")

	# Calculate Totals
	ws.cell(23,1).value = 'Total'
	col_sum = 0
	for col in range(2,coding_codons):
		for row in range(3,23):
			col_sum += ws.cell(row,col).value
			if row == 22:
				ws.cell(23,col).value = col_sum
				col_sum = 0
	total_count = 0
	for i in range(2,coding_codons):
		total_count += ws.cell(23,i).value
	ws.cell(23,coding_codons).value = total_count

	#Calculate Rates
	for col in range(2, coding_codons):
		for row in range(27, 47):
			if ws.cell(23,col).value != 0:
				ws.cell(row, col).value = ws.cell(row-row_offset,col).value/ws.cell(23,col).value

	ws.cell(50,1).value = 'muts'
	ws.cell(51,1).value = 'sites'
	ws.cell(52,1).value = 'rate'
	col_count = 2
	for key in aa_sites_dic:
		for codon in distance_lib[key]:
			ws.cell(49,col_count).value = codon
			codon_errors = codon_error_table[codon]
			ws.cell(50,col_count).value = codon_errors
			codon_wilds = codon_wild_table[codon]
			total_sites = codon_wilds + codon_errors
			ws.cell(51,col_count).value = total_sites
			if total_sites != 0:
				codon_rate = codon_errors / total_sites
			else:
				codon_rate = ''
			ws.cell(52,col_count).value = codon_rate
			col_count += 1

	wb.save(output)

if __name__ == '__main__':
	file_list = None
	try:
		options, remainder = getopt.getopt(sys.argv[1:], '', ['input_file=', 'gene_file=', 'protein_file=', 'tt='])
	except getopt.GetoptError:
		print('Example: python codon_matrix.py --input_file analyzed_processed_wt_ecoli-1_pep99.xlsx --gene_file wt_ecoli_genes.fasta --protein_file wt_ecoli_proteome.fasta --tt 11')
		sys.exit(2)

	input_file = None
	gene_file = None
	protein_file = None
	usage = None
	tt = None

	for opt, arg in options:
		if opt == '--input_file':
			input_file = arg
		elif opt == '--gene_file':
			gene_file = arg
		elif opt == '--protein_file':
			protein_file = arg
		elif opt == '--tt':
			tt = int(arg)

	if not input_file:
		print("Error: Missing input_file argument; the error analysis output file")
		sys.exit(2)
	elif not gene_file:
		print("Error: Missing gene_file argument")
		sys.exit(2)
	elif not protein_file:
		print("Error: Missing protein_file argument")
		sys.exit(2)
	elif not tt:
		print("Error: Missing translation table (tt) argument")
		sys.exit(2)

	codon_matrix(input_file, gene_file, protein_file, tt)
