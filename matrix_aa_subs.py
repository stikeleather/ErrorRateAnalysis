import re
import openpyxl
import sys
import getopt

def matrix_aa_subs(workbook_input):
	wb = openpyxl.Workbook()
	wb.create_sheet(index=0, title='Sheet1')
	file_name = re.split('[.]', workbook_input)
	workbook_out = f"matrix_{file_name[0]}.xlsx"
	wb.save(workbook_out)

	wb0 = openpyxl.load_workbook(workbook_input)
	ws0 = wb0["Sheet1"]

	row_num_count = 0
	for www in range(1,ws0.max_row+1):
		if ws0.cell(www,1).value is not None:
			row_num_count += 1
		else:
			break
	row_num = row_num_count+1

	wb = openpyxl.load_workbook(workbook_out)
	ws = wb["Sheet1"]

	split_name_dic = dict()
	seq_dic = dict()
	psm_count_dic = dict()
	key_count1 = 0
	dot_dash_pattern = re.compile(r'[-.]')
	for ww in range(2,row_num):
		key_count1 += 1
		name_to_split = ws0.cell(ww,1).value
		seq = ws0.cell(ww,2).value
		psm_count = ws0.cell(ww,3).value
		split_name = re.split(dot_dash_pattern, name_to_split)
		split_name_dic.update({key_count1:split_name})
		seq_dic.update({key_count1:seq})
		psm_count_dic.update({key_count1:psm_count})

	#Intiate two dictionaries; one for total sites and one for mutations
	aa_sites_dic = {
		'R':0,'H':0,'K':0,'D':0,'E':0,'S':0,'T':0,'N':0,'Q':0,'C':0,
		'G':0,'P':0,'A':0,'V':0,'I':0,'L':0,'M':0,'F':0,'Y':0,'W':0
		}
	aa_mut_dic = dict(aa_sites_dic)

	#Need integar values for worksheet operations; adds amino acid codes to output
	aa_codes = list(aa_sites_dic)
	for i in range(2, 22):
		ws.cell(1,i).value = aa_codes[i-2]
		ws.cell(i,1).value = aa_codes[i-2]
		ws.cell(24,i).value = aa_codes[i-2]
		ws.cell(i+23,1).value = aa_codes[i-2]
		ws.cell(47,i).value = aa_codes[i-2]


	#Intialize zeroes in matrix
	for q in range(2,22):
		for w in range(2,22):
			ws.cell(w,q).value = 0
			ws.cell(w+23,q).value = 0
	
	mut_pattern = re.compile(r'([A-Z])\d+([A-Z])$')
	#Loops over mutants; determines source aa and dest aa; finds the intersect in the matrix and adds the PSM count to the cell
	for key in split_name_dic:
		if split_name_dic[key][0] == 'mutant':
			src = re.match(mut_pattern, split_name_dic[key][4]).group(1)
			dest = re.match(mut_pattern ,split_name_dic[key][4]).group(2)
			if src not in aa_sites_dic.keys() or dest not in aa_sites_dic.keys():
				raise(f"Invalid source or destination: {key}, src: {src}, dest: {dest}")
			mut_amount = psm_count_dic[key]
			for w in range(2,22):
				for e in range(2,22):
					#Matrix intersect
					if ws.cell(1,w).value == src and ws.cell(e,1).value == dest:
						#Existing PSM + new PSM
						ws.cell(e,w).value = ws.cell(e,w).value + mut_amount


	col_sum = 0
	for q in range(2,22):
		for w in range(2,22):
			col_sum += ws.cell(w,q).value
			if w == 21:
				ws.cell(22,q).value = col_sum
				col_sum = 0
	ws.cell(22,1).value = 'Total'
	total_count = 0
	for i in range(2,22):
		total_count += ws.cell(22,i).value
	ws.cell(22,22).value = total_count


	for key2 in seq_dic:
		for i in seq_dic[key2]:
			aa = i
			for j in aa_sites_dic.keys():
				if j == aa:
					aa_sites_dic[j] += psm_count_dic[key2]
					break

				
	for key in seq_dic:
		if split_name_dic[key][0] == 'mutant':
			src = re.match(mut_pattern,split_name_dic[key][4]).group(1)
			for j in aa_sites_dic.keys():
				if j == src:
					aa_mut_dic[j] += psm_count_dic[key]
					aa_sites_dic[j] += psm_count_dic[key]
					aa_sites_dic[re.match(mut_pattern,split_name_dic[key][4]).group(2)] -= psm_count_dic[key]
					break


	for q in range(2,22):
		for key in aa_mut_dic:
			if ws.cell(1,q).value == key:
				ws.cell(48,q).value	= aa_mut_dic[key]
				ws.cell(49,q).value = aa_sites_dic[key]
				if aa_sites_dic[key] != 0:
					ws.cell(50,q).value = aa_mut_dic[key]/aa_sites_dic[key]
				else:
					ws.cell(50,q).value = 'Div/0'


	wb.save(workbook_out)


	ws.insert_rows(16)
	ws.cell(16,1).value = 'X'
	for rrr in range(2,22):
		ws.cell(16,rrr).value = ws.cell(17,rrr).value + ws.cell(18,rrr).value

	ws.delete_rows(17)
	ws.delete_rows(17)


	for i in range(2,22):
		aa_type = ws.cell(i,1).value
		ws.cell(i+22,1).value = aa_type

	for q in range(2,22):
		for w in range(2,22):
			numerator = ws.cell(w,q).value
			denominator = ws.cell(21,q).value
			if denominator != 0:
				ws.cell(w+22,q).value = numerator/denominator
			else:
				ws.cell(w+22,q).value = ''

	ws.delete_rows(43)

	ws.cell(21,1).value = 'Total'
	ws.cell(46,1). value = 'Muts'
	ws.cell(47,1). value = 'Sites'
	ws.cell(48,1). value = 'Rate'
	ws.cell(45,22).value = 'Total'


	mut_total = 0
	site_total = 0
	for ttt in range(2,22):
		mut_total += ws.cell(46,ttt).value
		site_total += ws.cell(47,ttt).value
		if ttt == 21:
			ws.cell(46,22).value = mut_total
			ws.cell(47,22).value = site_total
			ws.cell(48,22).value = mut_total/site_total

	total_sites = 0
	for key in seq_dic:
		total_sites += (len(seq_dic[key]))*psm_count_dic[key]

	total_muts = 0
	for key in split_name_dic:
		if split_name_dic[key][0] == 'mutant':
			total_muts += psm_count_dic[key]

	ws.cell(50,1).value = 'Total_sites'
	ws.cell(51,1).value = total_sites

	ws.cell(50,2).value = 'Total_muts'
	ws.cell(51,2).value = total_muts

	wb.save(workbook_out)
	return workbook_out

def print_usage():
    print("Usage:")
    print("Example: python matrix_aa_subs.py --workbook_input analyzed_processed_wt_ecoli-1_pep99.xlsx")
    print("Arguments:")
    print("--workbook_input : An analyzed file output from run_error_analysis")

if __name__ == '__main__':
    workbook_input = None

    try:
        options, remainder = getopt.getopt(sys.argv[1:], 'h', ['workbook_input=', 'help'])
    except getopt.GetoptError:
        print_usage()
        sys.exit(2)

    for opt, arg in options:
        if opt in ('-h', '--help'):
            print_usage()
            sys.exit(0)
        elif opt == '--workbook_input':
            workbook_input = arg
        else:
            print(f"Warning! Command-line argument: {opt} not recognized. Exiting...")
            sys.exit(2)

    if remainder:
        print(f"Unrecognized arguments: {remainder}. These were ignored.")

    if not workbook_input:
        print("Error: --workbook_input argument is required.")
        print_usage()
        sys.exit(2)

    matrix_aa_subs(workbook_input)

