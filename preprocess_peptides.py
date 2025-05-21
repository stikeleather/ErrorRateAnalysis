import getopt
import pickle
import re
import sys
import openpyxl
import pandas as pd
from Bio import SeqIO
from collections import defaultdict

def mutagenize(peptide):
	AA_dict = {'A':71.04,'R':156.10,'N':114.04,'D':115.03,'C':103.00,
	'E':129.04,'Q':128.06,'G':57.02,'H':137.06,'I':113.08,'L':113.08,
	'K':128.09,'M':131.04,'F':147.07,'P':97.05,'S':87.03,'T':101.04,
	'W':186.08,'Y':163.06,'V':99.07}
	mut_pep_list = list()
	for k in range(len(peptide)):
		for j in AA_dict:
			if j != peptide[k]:
				new_pep = peptide[:k] + j + peptide[(k+1):]
				mut_pep_list.append(new_pep)
	return mut_pep_list

def fasta_to_dict(fasta_file):
	sequence_dict = {}
	for record in SeqIO.parse(fasta_file, "fasta"):
		if record.id in sequence_dict.keys():
			print(f"Duplicate record id detected in fasta: {record.id}")
		else:
			sequence_dict[record.id] = str(record.seq)
	return sequence_dict

def find_duplicate_sequences_with_ids(fasta_input):
	seq_to_ids = defaultdict(list)
	if isinstance(fasta_input, dict):
		for header, seq in fasta_input.items():
			seq_to_ids[seq].append(header)
	else:
		for record in SeqIO.parse(fasta_input, "fasta"):
			seq_to_ids[str(record.seq)].append(record.id)
	return {seq: ids for seq, ids in seq_to_ids.items() if len(ids) > 1}

def find_duplicates_in_data(workbook_in, duplicate_seq_to_ids):
    if isinstance(workbook_in, pd.DataFrame):
        df = workbook_in.copy()
    else:
        df = pd.read_excel(workbook_in)

    duplicate_seqs = set(duplicate_seq_to_ids.keys())
    matched_df = df[df['Sequence'].isin(duplicate_seqs)]
    return matched_df

def find_match(long_pep, short_pep):
	try:
		start_index = long_pep.index(short_pep)
		end_index = start_index + len(short_pep)
		return start_index, end_index
	except ValueError:
		return -1, -1
	
def find_mismatch(wt_seq, mut_seq):
	mismatch_pos = -1
	match_pos = -1
	# Loop over each possible starting position in wt_seq where mut_seq can fit
	for i in range(len(wt_seq) - len(mut_seq) + 1):
		segment = wt_seq[i:i+len(mut_seq)]
		mismatch_indices = [j for j in range(len(mut_seq)) if segment[j] != mut_seq[j]]
		# Check for one mismatch
		if len(mismatch_indices) == 1:
			mismatch_pos = i + mismatch_indices[0]  # position of mismatch in wt_seq
			match_pos = i  # starting position of alignment in wt_seq
			return (match_pos + 1), (mismatch_pos + 1), (match_pos + len(mut_seq))
	return -1, -1

def resolve_duplicates(fasta_file, workbook_in, protein_dic, gene_dic, duplicate_seqs=None):
	duplicate_seqs_to_ids = duplicate_seqs or find_duplicate_sequences_with_ids(fasta_file)
	matched_data_duplicates = find_duplicates_in_data(workbook_in, duplicate_seqs_to_ids)
	matched_data_duplicates['accessions'] = matched_data_duplicates['Sequence'].map(duplicate_seqs_to_ids)
	dash_pattern = re.compile(r'[-]')
	disallowed_sequences = set()
	allowed_duplicates = set()
	allowed_sequences = set()
	for idx, row in matched_data_duplicates.iterrows():
		seq = row['Sequence']
		accessions = row['accessions']
		type_set = set()
		parent_set = set()
		for acc in accessions:
			dash_acc = re.split(dash_pattern, acc)
			id_type = dash_acc[0]
			parent_name = dash_acc[1]
			type_set.add(id_type)
			parent_set.add(parent_name)
		if parent_set.issubset(allowed_duplicates):
			continue
		if all((seq, parent) in allowed_sequences for parent in parent_set):
			continue		
		if "wild" in type_set and "mutant" not in type_set:
			dna_set = set()
			for parent in parent_set:
				gene_seq = gene_dic[parent]
				dna_set.add(gene_seq)
			if len(dna_set) == 1:
				allowed_duplicates.update(parent_set)
			else:
				dna_set = set()
				for parent in parent_set:
					protein_seq = protein_dic[parent]
					matches = find_match(protein_seq, seq)
					match_start = matches[0] * 3
					match_end = matches[1] * 3
					parent_gene_seq = gene_dic[parent]
					current_dna_segment = parent_gene_seq[match_start:match_end]
					dna_set.add(current_dna_segment)
				if len(dna_set) == 1:
					for parent in parent_set:
						allowed_sequences.add((seq, parent))
				else:
					disallowed_sequences.add(seq)
		elif "mutant" in type_set and "wild" not in type_set:
			dna_set = set()
			for parent in parent_set:
				gene_seq = gene_dic[parent]
				dna_set.add(gene_seq)
			if len(dna_set) == 1:
				allowed_duplicates.update(parent_set)
			else:
				dna_set = set()
				for parent in parent_set:
					protein_seq = protein_dic[parent]
					mismatches = find_mismatch(protein_seq, seq)
					match_start = (mismatches[0] * 3) - 3
					match_end = mismatches[1] * 3
					parent_gene_seq = gene_dic[parent]
					current_dna_segment = parent_gene_seq[match_start:match_end]
					dna_set.add(current_dna_segment)
					if len(dna_set) == 1:
						for parent in parent_set:
							allowed_sequences.add((seq, parent))
					else:
						disallowed_sequences.add(seq)
		else:
			disallowed_sequences.add(seq)

	# with open("disallowed_sequences.txt", 'a') as f:
	# 	for item in disallowed_sequences:
	# 		f.write(f"{item}\n")
	return disallowed_sequences


def preprocess_peptides(input_file, analyzed_workbook, homology_file, mut_fasta, gene_file, protein_file, xle_corr, duplicate_seqs):
	def remove_formatting(input_path):
		original_workbook = openpyxl.load_workbook(filename=input_path)

		new_workbook = openpyxl.Workbook()

		for sheet_name in original_workbook.sheetnames:
			original_sheet = original_workbook[sheet_name]
			new_sheet = new_workbook.create_sheet(title=sheet_name)

			for row in original_sheet.iter_rows(values_only=True):
				new_sheet.append(row)

		new_workbook.remove(new_workbook["Sheet"])
		new_workbook.save(filename=f"{input_path}")

	print("Removing Artifacts...")
	wb0 = openpyxl.load_workbook(input_file)
	ws0 = wb0["Sheet1"]

	df = pd.read_excel(input_file)

	col_names = dict()
	col_count = 0
	for col in range(1,ws0.max_column+1):
		if ws0.cell(1,col).value is not None:
			col_count += 1
			col_head = ws0.cell(1,col).value
			col_names.update({col_head:col_count})
		else:
			break

	seq = col_names['Sequence']
	psm = col_names['# PSMs']

	row_num0 = 0
	for www in range(1,ws0.max_row+1):
		if ws0.cell(www,psm).value is not None:
			row_num0 += 1
		else:
			break
	row_num0 = row_num0+1
	# Artifact regex patterns
	NtoD = re.compile(r'N\d+D')
	QtoE = re.compile(r'Q\d+E')
	EtoS = re.compile(r'E\d+S')
	StoD = re.compile(r'S\d+D')
	TtoE = re.compile(r'T\d+E')
	StoA = re.compile(r'S\d+A')
	YtoF = re.compile(r'Y\d+F')
	# Artifact Definitions
	ptm_definitions = {NtoD,QtoE,EtoS,StoD,TtoE,StoA,YtoF}

	wb = openpyxl.load_workbook(analyzed_workbook)
	ws = wb["Sheet1"]

	row_num = 0
	for www in range(1,ws.max_row+1):
		if ws.cell(www,1).value is not None:
			row_num += 1
		else:
			break
	row_num = row_num+1

	target_seq_set = set()
	for acc in range(2,row_num):
		if 'mutant-' in ws.cell(acc,1).value:
			accession = ws.cell(acc,1).value
			split_acc = re.split('[.]',accession)
			mutation = split_acc[2]
			for pattern in ptm_definitions:
				if pattern.search(mutation):
					target_seq = ws.cell(acc,2).value
					target_seq_set.add(target_seq)
					
	with open(homology_file, 'rb') as f:
		homolgy_peptides = pickle.load(f)

	for row in range(2,row_num):
		test = ws0.cell(row,seq).value
		if test in homolgy_peptides:
			target_seq_set.add(test)

	for item in homolgy_peptides:
		mut_peptides = mutagenize(item)
		for entry in mut_peptides:
			target_seq_set.add(entry)


	data_rows = []
	for index, row in df.iterrows():
		if pd.notnull(row['Sequence']) and row['Sequence'] != "":
			if row['Sequence'] not in target_seq_set:
				data_rows.append(row)

	processed_data = pd.DataFrame(data_rows, columns=df.columns)

	if isinstance(protein_file, dict):
		protein_dic = protein_file
	else:
		protein_dic = fasta_to_dict(protein_file)

	if isinstance(gene_file, dict):
		gene_dic = gene_file
	else:
		gene_dic = fasta_to_dict(gene_file)

	disallowed_sequences = resolve_duplicates(mut_fasta, df, protein_dic, gene_dic, duplicate_seqs)
	processed_data = processed_data[~processed_data['Sequence'].isin(disallowed_sequences)]

	with open(xle_corr,'rb') as f:
		xle_correction = pickle.load(f)
	xle_keys = set(xle_correction.keys())
	for idx, row in df.iterrows():
		raw_seq = row['Sequence']
		if raw_seq in xle_keys:
			df.loc[idx, 'Sequence'] = xle_correction[raw_seq]

	disallowed_sequences = resolve_duplicates(mut_fasta, df, protein_dic, gene_dic, duplicate_seqs)
	processed_data = processed_data[~processed_data['Sequence'].isin(disallowed_sequences)]


	out_file = f"processed_{input_file}"
	processed_data.to_excel(out_file,index=False)
	remove_formatting(out_file)
	return out_file
	
if __name__ == '__main__':
	try:
		options, remainder = getopt.getopt(sys.argv[1:],'', ['input_file=','analyzed_workbook=', 'homology_file=', 'mut_fasta=', 'gene_file=', 'protein_file=', 'xle_corr=', 'duplicate_seqs='])
	except getopt.GetoptError:
		print("Example: python preprocess_peptides.py --input_file wt_ecoli-1_pep99.xlsx --analyzed_workbook analyzed_wt_ecoli-1_pep99.xlsx --homology_file homology_ih_mut_custom_wt_ecoli_proteome.pkl")

	for opt, arg in options:
		if opt == '--input_file': 
			input_file=arg
		elif opt == '--analyzed_workbook': 
			analyzed_workbook=arg
		elif opt == '--homology_file': 
			homology_file = arg
		elif opt == "--mut_fasta":
			mut_fasta = arg
		elif opt == '--gene_file':
			gene_file = arg
		elif opt == '--protein_file':
			protein_file = arg
		elif opt == '--xle_corr':
			xle_corr = arg
		elif opt == '--duplicate_seqs':
			duplicate_seqs = arg								
		else: 
			print("Warning! Command-line argument: %s not recognized. Exiting..." % opt); sys.exit(2)

	if remainder:
		print(f"Unrecognized arguments: {remainder}. These were ignored.")

	preprocess_peptides(input_file, analyzed_workbook, homology_file, mut_fasta, gene_file, protein_file, xle_corr, duplicate_seqs)