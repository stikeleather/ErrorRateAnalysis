import getopt
import re
import sys
import time
import openpyxl
from Bio import SeqIO
import pickle
from collections import defaultdict
from matrix_aa_subs import matrix_aa_subs
from generate_custom_database import fasta_to_dict
from preprocess_peptides import preprocess_peptides
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment
from codon_bias import codon_bias
from calculate_stats import calculate_stats
from format_matrix import format_matrix
from codon_matrix import codon_matrix


def find_match(long_pep, short_pep):
	try:
		start_index = long_pep.index(short_pep)
		end_index = start_index + len(short_pep)
		return start_index, end_index
	except ValueError:
		return -1, -1
	
def find_mismatch(wt_seq, mut_seq):
	mismatch_pos = -1
	match_pos = -1
	# Loop over each possible starting position in wt_seq where mut_seq can fit
	for i in range(len(wt_seq) - len(mut_seq) + 1):
		segment = wt_seq[i:i+len(mut_seq)]
		mismatch_indices = [j for j in range(len(mut_seq)) if segment[j] != mut_seq[j]]
		# Check for one mismatch
		if len(mismatch_indices) == 1:
			mismatch_pos = i + mismatch_indices[0]  # position of mismatch in wt_seq
			match_pos = i  # starting position of alignment in wt_seq
			return (match_pos + 1), (mismatch_pos + 1), (match_pos + len(mut_seq))
	return -1, -1

def determine_sub_pos(long_pep, short_pep, mutation):
	start_range = long_pep.find(short_pep)
	pos = int(mutation.group(2))
	sub_pos = pos - start_range
	new_seq = short_pep[:sub_pos-1] + mutation.group(1) + short_pep[sub_pos:]
	return new_seq, sub_pos

def count_mismatches(test, comp):
	if len(test) != len(comp):
		raise ValueError("Test and Comp must be equal in length")
	mismatch_count = 0
	for aa1, aa2 in zip(test, comp):
		if aa1 != aa2:
			mismatch_count += 1
	return mismatch_count

def read_file_names(file_input):
	if isinstance(file_input, list):
		return file_input
	else:
		with open(file_input, 'r') as file:
			file_names = file.read().splitlines()
		return file_names

def extract_species_name(filename):
	species_pattern = r'ih_mut_custom_(.*?)_proteome\.fasta'
	match = re.search(species_pattern, filename)
	if match:
		return match.group(1)
	else:
		return None

def filter_pep(input_file):
	data = pd.read_excel(input_file)
	data = data[data["PEP"] < 0.01]
	file_name = re.split("[.]",input_file)
	file_name = file_name[0]
	output_file = f"{file_name}_pep99.xlsx"
	data.to_excel(output_file, index=False)
	return output_file

def find_duplicate_sequences_with_ids(fasta_input):
	seq_to_ids = defaultdict(list)
	if isinstance(fasta_input, dict):
		for header, seq in fasta_input.items():
			seq_to_ids[seq].append(header)
	else:
		for record in SeqIO.parse(fasta_input, "fasta"):
			seq_to_ids[str(record.seq)].append(record.id)
	return {seq: ids for seq, ids in seq_to_ids.items() if len(ids) > 1}

def update_nsps_seqs(mut_fasta, nsp_seqs, written_nsps):
	if len(nsp_seqs) == 0:
		return
	print("Writing NSPs to File")
	with open(mut_fasta, 'a') as f:
		for header, sequence in nsp_seqs.items():
			if header not in written_nsps:
				f.write(f">{header}\n")
				f.write(f"{sequence}\n")
				written_nsps.add(header)
	return written_nsps


def resolve_wild_nsp_parent_peptide(nsp_peptide, current_parent_name, test_parent_name, protein_dictionary, gene_dictionary, different_dna_set):
	if nsp_peptide in different_dna_set:
		nsp_status = False
		return nsp_status, different_dna_set
	else:
		nsp_status = True
		current_parent_nt_sequence = gene_dictionary[current_parent_name]
		test_parent_protein_seq = protein_dictionary[test_parent_name]
		matches = find_match(test_parent_protein_seq, nsp_peptide)
		test_parent_nt_sequence = gene_dictionary[test_parent_name]
		match_start = matches[0] * 3
		match_end = matches[1] * 3
		current_dna_segment = current_parent_nt_sequence[match_start:match_end]
		test_dna_segment = test_parent_nt_sequence[match_start:match_end]
		if current_dna_segment != test_dna_segment:
			different_dna_set.add(nsp_peptide)
			nsp_status = False
			#True if no different DNA is detected, else false
			return nsp_status, different_dna_set
		return nsp_status, different_dna_set

def resolve_mutant_nsp_parent_peptide(nsp_peptide, current_parent_name, test_parent_name, protein_dictionary, gene_dictionary, different_dna_set):
	if nsp_peptide in different_dna_set:
		nsp_status = False
		return nsp_status, different_dna_set
	else:
		nsp_status = True
		current_parent_nt_sequence = gene_dictionary[current_parent_name]
		test_parent_protein_seq = protein_dictionary[test_parent_name]
		mismatches = find_mismatch(test_parent_protein_seq, nsp_peptide)
		test_parent_nt_sequence = gene_dictionary[test_parent_name]
		match_start = (mismatches[0] * 3) - 3
		match_end = mismatches[1] * 3
		current_dna_segment = current_parent_nt_sequence[match_start:match_end]
		test_dna_segment = test_parent_nt_sequence[match_start:match_end]
		if current_dna_segment != test_dna_segment:
			different_dna_set.add(nsp_peptide)
			nsp_status = False
			#True if no different DNA is detected, else false
			return nsp_status, different_dna_set
		return nsp_status, different_dna_set

def run_error_analysis(mut_fasta, workbook_input, protein_file, gene_file, xle_seqs=None, xle_corr=None, nsp_dic=None, wild_seqs_dic=None, mutant_seqs_dic=None, different_dna_set=None):
	#Removes peptides with PEP scores greater than 0.01
	workbook_out = f'analyzed_{workbook_input}'
	#Output xlsx creation
	wb = openpyxl.Workbook()
	wb.create_sheet(index=0, title='Sheet1')
	wb.remove(wb["Sheet"])
	wb.save(workbook_out)
	#Set worksheet-0 to the raw data
	#Raw data is never saved here, so any edits are temporary; integrity of raw data is secure

	if isinstance(different_dna_set, str):
		with open(different_dna_set, 'rb') as f:
			different_dna_set = pickle.load(f)

	if isinstance(different_dna_set, set) and different_dna_set:
		df = pd.read_excel(workbook_input)
		df = df[~df["Sequence"].isin(different_dna_set)].copy()
		df.to_excel(workbook_input, index=False)

	wb0 = openpyxl.load_workbook(workbook_input)
	ws0 = wb0["Sheet1"]

	#Determines the number of rows; avoids calling ws0.max_row()
	row_num = 0
	for www in range(1,ws0.max_row+2):
		if ws0.cell(www,1).value is not None:
			row_num += 1
		else:
			break
	row_num = row_num+1

	#Dictionary of column names; makes it so order of columns does not matter
	col_names = dict()
	col_count = 0
	for col in range(1,ws0.max_column+1):
		if ws0.cell(1,col).value is not None:
			col_count += 1
			col_head = ws0.cell(1,col).value
			col_names.update({col_head:col_count})
		else:
			break

	#Sets column integars by column name
	mpa = col_names['Master Protein Accessions']
	seq_col = col_names['Sequence']
	psm = col_names['# PSMs']

	#If empty cells are present in the accession column, add NA-count string
	na_count = 0
	total_count = 0
	for v in range(2, row_num):
		total_count += 1
		if ws0.cell(v,mpa).value is None:
			na_count += 1
			ws0.cell(v,mpa).value = f"NA-{na_count}"

	#Dictionary of peptides in the data
	search_dic = dict()
	for row in range(2,row_num):
		if ws0.cell(row,seq_col).value is not None:
			search_dic.update({row:ws0.cell(row,seq_col).value})
		else:
			break

	#Intiate Regular Expression patterns
	ItoL = re.compile(r'I\d+L')
	LtoI = re.compile(r'L\d+I')
	# The $ anchors mut_patten to the end of the string, avoiding issues with locus ids
	mut_pattern = re.compile(r'([A-Z])(\d+)([A-Z])$')
	dash_pattern = re.compile(r'[-]')
	dot_pattern = re.compile(r'[.]')
	dash_dot_pattern = re.compile(r'[-.]')
	header_pattern = re.compile(r'mutant-([^\.]+)')

	#Reads the mut_fasta either from a file for the first pass, or uses the one in mem; creates dictionary of headers:seqs
	mutant_fasta = mut_fasta
	if not wild_seqs_dic and not mutant_seqs_dic:
		print("Sorting peptide database...")
		for key in mutant_fasta:
			split_key = re.split(dash_pattern,key)
			if split_key[0] == 'wild':
				wild_seqs_dic[key] = mutant_fasta[key]
			else:
				mutant_seqs_dic[key] = mutant_fasta[key]

	nsps_to_pop = set()
	if nsp_dic:
		mutant_fasta.update(nsp_dic)
		for key in nsp_dic:
			split_key = re.split(dash_pattern,key)
			if split_key[0] == 'wild':
				wild_seqs_dic[key] = nsp_dic[key]
			else:
				mutant_seqs_dic[key] = nsp_dic[key]

	print("Correcting for Xle...")
	#Loops through all possible Xle mutations; compares to known sequences in data; reverts to wt seq if Xle sub is found;
	#Stores raw_seq:xle_correction_seq in dict; will dump dic to pkl for future loading (makes data reruns faster); not suitable between samples
	#i.e. do not use xle correction for sample-1 on sample-2 for individual sample runs; leave xle_corr empty; Or just use the file_list option
	xle_dump = False
	xle_correction = dict()
	if not xle_corr:
		for key in xle_seqs:
			target_seq = xle_seqs[key]
			for entry in search_dic:
				if search_dic[entry] == target_seq:
					header_match = re.search(header_pattern,key)
					header_match = header_match.group(1)
					wild_header = f"wild-{header_match}"
					wild_seq = mutant_fasta[wild_header]
					# print(f"{key}")
					# print(target_seq)
					# print(wild_header)
					# print(wild_seq)
					ws0.cell(entry,seq_col).value = wild_seq
					xle_correction[target_seq] = wild_seq
					xle_dump = True

		split_workbook = re.split(dot_pattern,workbook_input)
		workbook_name = split_workbook[0]
		xle_corr = f'xle_{workbook_name}.pkl'

	else:
		with open(xle_corr,'rb') as f:
			xle_correction = pickle.load(f)
			xle_keys = set(xle_correction.keys())
			print("Loaded Xle Correction")
			for entry in search_dic:
				if search_dic[entry] in xle_keys:
					raw_seq = search_dic[entry]
					ws0.cell(entry,seq_col).value = xle_correction[raw_seq]

	#Assigning output file column names; order is important
	wb = openpyxl.load_workbook(workbook_out)
	ws = wb["Sheet1"]
	column_names = ['accession', 'Sequence', 'PSM_count', 'pep_length', 'wild_aa_total', 'mutant_aa_total', 'total_aa', 'error_rate']
	for i in range(1,9):
		ws.cell(1,i).value = column_names[i-1]

	#Dictionary where dic[header] = seq is flipped to dic[seq] = header
	print("Building Dictionaries")
	name_seq_dic = dict()
	for name, seq in mutant_fasta.items():
		name_seq_dic[seq] = name


	#First rebuild of the search_dic (post xle correction)
	print("Building Search Targets")
	search_dic = dict()
	for row in range(2,row_num):
		search_dic.update({row:ws0.cell(row,seq_col).value})
	#Set of observed sequences
	observed_sequences = set(search_dic.values())
	final_dic_len = (len(observed_sequences))


	# name_seq_dic.keys() are peptide sequences; creates a set of peptide sequences for intersection with set of observed sequences
	#Non-standard peptide products will miss this intersection
	unique_name_seq_terms = set(name_seq_dic.keys())
	exist_set = observed_sequences.intersection(unique_name_seq_terms)


	#Compares how many sequences intersected to the number of observed sequences
	print(f"Matched {len(exist_set)} of {final_dic_len} unique peptides")

	#Determines how many non-standard sequences are in the observed data
	total_unknowns = 0
	test_terms = set()
	for i in range(2,row_num):
		test = ws0.cell(i,seq_col).value
		if test not in exist_set:
			test_terms.add(test)
	total_unknowns = len(test_terms)

	#Determines nsp id starting value to ensure previous nsp ids are not overwritten
	nsp_start_number = 0
	for key in mutant_fasta:
		if "-nsp" in key:
			nsp_start_number += 1

	nsps_for_filtering = set()
	if total_unknowns > 0:
		print(f"{total_unknowns} Non-standard products detected")
		unknown_count = 0
		total_uknown_wilds = 0
		total_uknown_muts = 0
		#First checks to see if the test seq is found in any wild-type peptide (comparison peptide)
		#Only comparison peptides used are those found to be in the data
		#If any matches are found then peptide is checked for homologous matches
		#If a peptide maps to more than 1 protein, then if the DNA is the same, then it is kept
		#Otherwise, if the DNA is different, then the NSP is filtered
		#note, we are looping over a copy() of test_terms and discarding from actual test terms
		#Each NSP needs a unique key
		print("Testing for wild NSPs...")
		for test in test_terms.copy():
			nsp_status = True
			current_parent_id = None
			current_parent_name = None
			for comp_id in wild_seqs_dic:
				comp_seq = wild_seqs_dic[comp_id]
				if test in comp_seq and "-nsp" not in comp_id:
					split_current_parent = re.split(dash_pattern, comp_id)
					test_parent_name = split_current_parent[1]
					if current_parent_name is None:
						current_parent_name = test_parent_name
					nsp_homolgy_check = resolve_wild_nsp_parent_peptide(test, current_parent_name, test_parent_name, protein_file, gene_file, different_dna_set)
					this_nsp_status = nsp_homolgy_check[0]
					if not this_nsp_status:
						nsp_status = False
						current_parent_id = None
						nsps_for_filtering.add(test)
						test_terms.discard(test)
						different_dna_set = nsp_homolgy_check[1]
						break
					elif this_nsp_status and current_parent_id is None:
						current_parent_id = comp_id
						split_current_parent_id = re.split(dash_pattern, current_parent_id)
						current_parent_name = split_current_parent_id[1]
			if nsp_status and current_parent_id is not None:
				unknown_count += 1
				total_uknown_wilds += 1
				print(f"Matched {unknown_count} of {total_unknowns} NSPs")
				wild_nsp_name = f"{current_parent_id}-nsp{nsp_start_number+unknown_count}"
				# Wild nsp will be added to database file
				nsp_dic[f"{wild_nsp_name}"] = f"{test}"
				# Wild nsp is added to the mutant fasta in memory
				mutant_fasta[wild_nsp_name] = test
				nsps_to_pop.add(wild_nsp_name)
				test_terms.discard(test)
		#If test terms still exist after searching for wilds, then search for mutants
		print("Testing for mutant NSPs...")
		if len(test_terms) > 0:
			for unk in test_terms.copy():
				nsp_status = True
				current_parent_id = None
				parent_split_name = None
				current_parent_seq = None
				parent_mutation = None
				current_parent_name = None
				parent_correction = None
				for comp_id in mutant_seqs_dic:
					comp_seq = mutant_seqs_dic[comp_id]
					if unk in comp_seq and "-nsp" not in comp_id:
						split_name = re.split(dash_dot_pattern, comp_id)
						test_parent_name = split_name[1]
						mutation = re.search(mut_pattern, comp_id)
						if current_parent_name is None:
							current_parent_name = test_parent_name
						#If the nsp is Xle, then correct it to wild-type and determine sub position
						if re.search(ItoL, comp_id) or re.search(LtoI, comp_id):
							correction = determine_sub_pos(comp_seq, unk, mutation)
							new_seq = correction[0]
							nsp_homolgy_check = resolve_wild_nsp_parent_peptide(new_seq, current_parent_name, test_parent_name, protein_file, gene_file, different_dna_set)
							this_nsp_status = nsp_homolgy_check[0]
							if not this_nsp_status:
								nsp_status = False
								current_parent_id = None
								nsps_for_filtering.add(unk)
								test_terms.discard(unk)
								different_dna_set = nsp_homolgy_check[1]
								break
							elif this_nsp_status and current_parent_id is None:
								current_parent_id = comp_id
								parent_split_name  = split_name
								current_parent_seq = comp_seq
								parent_mutation = mutation
								split_current_parent_id = re.split(dash_pattern, current_parent_id)
								current_parent_name = split_current_parent_id[1]

						else:
							nsp_homolgy_check = resolve_mutant_nsp_parent_peptide(unk, current_parent_name, test_parent_name, protein_file, gene_file, different_dna_set)
							this_nsp_status = nsp_homolgy_check[0]
							if not this_nsp_status:
								nsp_status = False
								current_parent_id = None
								nsps_for_filtering.add(unk)
								test_terms.discard(unk)
								different_dna_set = nsp_homolgy_check[1]
								break
							elif this_nsp_status and current_parent_id is None:
								current_parent_id = comp_id
								parent_split_name = split_name
								current_parent_seq = comp_seq
								parent_mutation = mutation
								split_current_parent_id = re.split(dash_pattern, current_parent_id)
								current_parent_name = split_current_parent_id[1]								
				if nsp_status and current_parent_id is not None:
					unknown_count += 1
					print(f"Matched {unknown_count} of {total_unknowns} NSPs")
					parent_correction = determine_sub_pos(current_parent_seq,unk,parent_mutation)
					new_seq = parent_correction[0]
					sub_pos = parent_correction[1]
					if re.search(ItoL, current_parent_id) or re.search(LtoI, current_parent_id):
						#Replaces an Xle sequence with the corrected wt seq in the current dataframe
						#The raw data workbook is never saved after editing; any edits are not permanent
						for entry in search_dic:
							if unk == search_dic[entry]:
								ws0.cell(entry,seq_col).value = new_seq
						xle_correction[unk] = new_seq
						xle_dump = True
						total_uknown_wilds += 1
						mutant_nsp_name = f"mutant-{parent_split_name[1]}-{parent_split_name[2]}.{parent_split_name[3]}.{parent_mutation.group(1)}{sub_pos}{parent_mutation.group(3)}-nsp{nsp_start_number+unknown_count}"
						#Xle NSP saved as mutant to identify others with same xle in next iteration; xle correction reverts these to wild-types
						nsp_dic[mutant_nsp_name] = unk
						#mutant fasta database is updated during this iteration to identify this xle nsp as wild-type
						wild_nsp_name = f"wild-{parent_split_name[1]}-{parent_split_name[2]}-nsp{nsp_start_number+unknown_count}"
						mutant_fasta[wild_nsp_name] = new_seq
						#ID stored for removal
						nsps_to_pop.add(wild_nsp_name)
					else:
						total_uknown_muts += 1
						#If nsp is not xle, then determine sub position for correct naming
						#Correct naming is important for usage of NSPs in downstream codon_bias
						mutant_nsp_name = f"mutant-{parent_split_name[1]}-{parent_split_name[2]}.{parent_split_name[3]}.{parent_mutation.group(1)}{sub_pos}{parent_mutation.group(3)}-nsp{nsp_start_number+unknown_count}"
						#mutant nsp will be added to database file
						nsp_dic[mutant_nsp_name] = unk
						#mutant nsp stored to memory
						mutant_fasta[mutant_nsp_name] = unk
						#ID stored for removal
						nsps_to_pop.add(mutant_nsp_name)
					test_terms.discard(unk)


		#Determines likely homology peptides that appeared in the data
		homology_discard = False
		if len(test_terms) > 0:
			discarded_homology = set()
			with open(homology_file, 'rb') as f:
				homolgy_peptides = pickle.load(f)
			for test in test_terms.copy():
				test_len = len(test)
				for comp in homolgy_peptides:
					if test_len == len(comp):
						num_mismatches = count_mismatches(test,comp)
						if num_mismatches == 1 or num_mismatches == 0:
							test_terms.discard(test)
							discarded_homology.add(test)
							homology_discard = True
							break
			if homology_discard:
				print("The following peptides were ignored due to internal homology:")
				print(discarded_homology)
		#Anything still remaining
		if len(test_terms) > 0:
			print("Unmatched NSPs:")
			print(test_terms)

		total_filtered_nsps = 0
		for v in range(2, row_num):
			peptide = ws0.cell(v, seq_col).value
			if peptide in nsps_for_filtering:
				total_filtered_nsps += 1	

		print(f"Found {total_unknowns} non-standard products")
		if total_unknowns > 0 and total_filtered_nsps > 0:

			if total_uknown_wilds > 0:
				print(f"{total_uknown_wilds} non-standard wilds")

			if total_uknown_muts > 0:
				print(f"{total_uknown_muts} non-standard muts")

			print(f"{len(nsps_for_filtering)} unique non-standard products filtered")

			print(f"{total_filtered_nsps} total NSPs were filtered due to homology")


		elif total_unknowns > 0 and total_filtered_nsps == 0:
			print(f"{total_uknown_wilds} non-standard wilds")
			print(f"{total_uknown_muts} non-standard muts")

		print("Re-Building Search Targets")
		# Collects raw peptide sequences from data input; including updated xle seqs stored in the copy of the xlsx
		search_dic = dict()
		for zz in range(2,row_num):
			search_dic.update({zz:ws0.cell(zz,seq_col).value})

		#Determines number of unique sequences, and creates set of sequences in raw data
		observed_sequences = set(search_dic.values())
		if nsps_for_filtering:
			observed_sequences -= nsps_for_filtering
		final_dic_len = (len(observed_sequences))

		#Dictionary where sequences are keys, and names are values
		name_seq_dic = {}
		for name, seq in mutant_fasta.items():
			name_seq_dic[seq] = name

		#Creates a set of all sequences to hashtable intersect observed peptide sequences
		#Results in a set of sequences
		unique_name_seq_terms = set(name_seq_dic.keys())
		exist_set = observed_sequences.intersection(unique_name_seq_terms)

		#Each sequence should be matched to a name
		psm_dic = dict()
		for term in exist_set:
			psm_dic.update({term:name_seq_dic[term]})
		print(f"Matched {len(exist_set)} of {final_dic_len} unique peptides")
	else:
		psm_dic = dict()
		for term in exist_set:
			psm_dic.update({term:name_seq_dic[term]})

	print("Matching Names to PSMs...")
	row_count2 = 2
	running_count = 0
	match_count = 0
	for key in psm_dic:
		running_count = 0
		for z in range(2,row_num):
			#print(ws0.cell(z,col_names['Sequence']).value)
			if key == ws0.cell(z,col_names['Sequence']).value:
				ws0.cell(z,mpa).value = psm_dic[key]
				running_count += ws0.cell(z,psm).value
				match_count += 1

		if running_count > 0:
			ws.cell(row_count2,1).value = psm_dic[key]
			ws.cell(row_count2,2).value = key
			ws.cell(row_count2, 3).value = running_count
			ws.cell(row_count2, 4).value = len(key)
			row_count2 += 1
			running_count = 0

	print(f"Matched {match_count} of {total_count} PSMs to names")


	print("Updating Database...")
	# Removes non-FASTA formatted NSP keys from memory; will be restored with FASTA-formatted ones in next iteration
	for key in nsps_to_pop:
		mutant_fasta.pop(key)

	#Math
	wb.save(workbook_out)
	wild_aa = 0
	mut_aa = 0
	total_aa = 0
	for row2 in range(2,ws.max_row+1):
		if ws.cell(row2,1).value is not None:
			acc = ws.cell(row2,1).value
			split_acc = re.split(dash_pattern, acc)
			if split_acc[0] == 'wild':
				psm_value = ws.cell(row2,3).value
				pep_len = ws.cell(row2,4).value
				wild_aa_result = psm_value * pep_len
				wild_aa += wild_aa_result
				total_aa += wild_aa_result
			else:
				psm_value = ws.cell(row2,3).value
				pep_len = ws.cell(row2,4).value
				wild_aa += (psm_value * pep_len) - psm_value
				mut_aa += psm_value
				total_aa += psm_value * pep_len
		else:
			break

	ws.cell(2,5).value = wild_aa
	ws.cell(2,6).value = mut_aa
	ws.cell(2,7).value = total_aa
	ws.cell(2,8).value = mut_aa/total_aa

	#Dumps the xle corrections made; can be used between runs of the same data input
	if xle_dump:
		print("Writing Xle correction...")
		with open(xle_corr,'wb') as q:
			pickle.dump(xle_correction,q)
	#Final workbook save
	wb.save(workbook_out)

	print("Generating Matrix...")
	matrix_out = matrix_aa_subs(workbook_out)

	return workbook_input, workbook_out, xle_corr, matrix_out, mutant_fasta, nsp_dic, wild_seqs_dic, mutant_seqs_dic, different_dna_set



def print_usage():
	print("Usage:")
	print("Example: python run_error_analysis.py --mut_fasta ih_mut_custom_example_proteome.fasta --workbook_input example.xlsx --xle_corr xle.pkl --gene_file genes.fasta --protein_file proteome.fasta --tt 11 --usage abs")
	print("Minimum 1: python run_error_analysis.py --mut_fasta ih_mut_custom_example_proteome.fasta --workbook_input example.xlsx --gene_file genes.fasta --protein_file proteome.fasta --tt 11 --usage abs")
	print("Minimum 2: python run_error_analysis.py --mut_fasta ih_mut_custom_example_proteome.fasta --file_list example_file_list.txt --gene_file genes.fasta --protein_file proteome.fasta --tt 11 --usage abs")
	print("\nArguments:")
	print("--mut_fasta     : Mutant FASTA file, made from generate_custom_database")
	print("--workbook_input: Exported peptide groups tab from Proteome Discoverer, .xlsx format  (optional, but required if no --file_list is provided)")
	print("--file_list     : Text file containing a list of workbook_input files (optional, but required if no --workbook_input is provided)")
	print("--protein_file  : FASTA file containing protein sequences from parse_genbank")
	print("--gene_file     : FASTA file containing gene sequences from parse_genbank")
	print("--tt            : Translation table number (e.g., 11) ")
	print("--usage         : Usage type, either 'abs' (absolute) or 'rel' (relative)")
	print("--xle_corr      : Optional pickle file to load xle correction file (.pkl)")
	print("--nsp_homology  : Optional pickle file to load NSP peptide sequences that are known to be homologous between sequences with different underlying DNA")

if __name__ == '__main__':
	xle_seqs_mem = None
	file_list = None
	combined_data = pd.DataFrame()
	mut_fasta = None
	workbook_input = None
	gene_file = None
	protein_file = None
	usage = None
	tt = None
	mutant_fasta_memory = None
	protein_dic_mem = None
	gene_dic_mem = None
	wild_seqs_memory = dict()
	mutant_seqs_memory = dict()
	nsp_dic = dict()
	different_dna_set = set()
	written_nsps = set()

	try:
		options, remainder = getopt.getopt(sys.argv[1:], 'h', ['mut_fasta=', 'workbook_input=', 'xle_corr=', 'file_list=', 'gene_file=', 'protein_file=', 'tt=', 'usage=', 'nsp_homology=', 'help'])
	except getopt.GetoptError:
		print_usage()
		sys.exit(2)

	for opt, arg in options:
		if opt in ('-h', '--help'):
			print_usage()
			sys.exit(0)
		elif opt == '--mut_fasta':
			mut_fasta = arg
		elif opt == '--workbook_input':
			workbook_input = str(arg)
		elif opt == '--xle_corr':
			xle_corr = arg
		elif opt == '--file_list':
			file_list = arg
		elif opt == '--gene_file':
			gene_file = arg
		elif opt == '--protein_file':
			protein_file = arg
		elif opt == '--tt':
			tt = int(arg)
		elif opt == '--usage':
			usage = arg
		elif opt == '--nsp_homology':
			different_dna_set = arg			
		else:
			print(f"Warning! Command-line argument: {opt} not recognized. Exiting...")
			sys.exit(2)

	if remainder:
		print(f"Unrecognized arguments: {remainder}. These were ignored.")

	if usage not in ['abs', 'rel']:
		print("Error: The 'usage' argument must be either 'abs' or 'rel'.")
		sys.exit(2)

	if not mut_fasta:
		print("Error: Missing mut_fasta argument")
		sys.exit(2)
	if not workbook_input and not file_list:
		print("Error: Missing workbook_input or file_list argument")
		sys.exit(2)
	if not gene_file:
		print("Error: Missing gene_file argument")
		sys.exit(2)
	if not protein_file:
		print("Error: Missing protein_file argument")
		sys.exit(2)
	if not tt:
		print("Error: Missing translation table (tt) argument")
		sys.exit(2)

	if file_list:
		workbook_inputs = read_file_names(file_list)
		filtered_file_list = []
		codon_file_list = []
	else:
		workbook_inputs = [workbook_input]

	file_count = 0
	total_files = len(workbook_inputs)
	start_time_all = time.time()
	print("Reading input file...")
	mutant_fasta_memory = mutant_fasta_memory or fasta_to_dict(mut_fasta)
	print("Preparing homology search...")
	duplicate_sequences = find_duplicate_sequences_with_ids(mutant_fasta_memory)
	for workbook_input in workbook_inputs:
		xle_corr = None
		file_count += 1
		start_time = time.time()
		workbook_input = filter_pep(workbook_input)
		if file_count == 1:
			#Sets homology file
			split_mut_fasta = re.split(r'[.]',mut_fasta)
			homology_file = f"homology_{split_mut_fasta[0]}.pkl"
			split_proteome_input = re.split(r'[.]', protein_file)
			xle_seqs_file = f'xle_seqs_custom_{split_proteome_input[0]}.pkl'
			with open(xle_seqs_file, 'rb') as f:
				xle_seqs_mem = pickle.load(f)
		print(f"Processing File {file_count} of {total_files}")
		mutant_fasta_memory = mutant_fasta_memory or fasta_to_dict(mut_fasta)
		protein_dic_mem = protein_dic_mem or fasta_to_dict(protein_file)
		gene_dic_mem = gene_dic_mem or fasta_to_dict(gene_file)
		out_items = run_error_analysis(mutant_fasta_memory, workbook_input, protein_dic_mem, gene_dic_mem, xle_seqs_mem, xle_corr, nsp_dic, wild_seqs_memory, mutant_seqs_memory, different_dna_set)
		workbook_input = out_items[0]
		workbook_out = out_items[1]
		xle_corr = out_items[2]
		mutant_fasta_memory = out_items[4]
		nsp_dic = out_items[5]
		wild_seqs_memory = out_items[6]
		mutant_seqs_memory = out_items[7]
		different_dna_set = out_items[8]
		if nsp_dic:
			written_nsps = update_nsps_seqs(mut_fasta, nsp_dic, written_nsps)
		processed_file = preprocess_peptides(workbook_input, workbook_out, homology_file, mutant_fasta_memory, gene_dic_mem, protein_dic_mem, xle_corr, duplicate_sequences)
		if file_list:
			filtered_file_list.append(processed_file)
		out_items = run_error_analysis(mutant_fasta_memory, processed_file, protein_dic_mem, gene_dic_mem, xle_seqs_mem, xle_corr, nsp_dic, wild_seqs_memory, mutant_seqs_memory, different_dna_set)
		workbook_out = out_items[1]
		mutant_fasta_memory = out_items[4]
		wild_seqs_memory = out_items[6]
		mutant_seqs_memory = out_items[7]
		different_dna_set = out_items[8]
		format_matrix(out_items[3], remove_artifacts=True)
		codon_output_file = codon_bias(workbook_out, gene_file, protein_file, tt, usage, mutant_fasta_memory)
		codon_matrix(workbook_out, gene_file, protein_file, tt)
		if file_list:
			codon_file_list.append(codon_output_file)

		print("Finished")
		end_time = time.time()
		total_time = end_time - start_time
		if total_time < 60:
			print(f"Total time for file {file_count} of {total_files}: {total_time:.2f} seconds")
		else:
			minutes = int(total_time // 60)
			seconds = int(total_time % 60)
			print(f"Total time for file {file_count} of {total_files}: {minutes} minute(s) {seconds} second(s)")


	if file_list:
		start_time = time.time()
		print("Processing all data...")
		species_name = extract_species_name(mut_fasta)
		nsp_homology_file = f"{species_name}_nsp_homology.pkl"
		stats_output_file = f'{species_name}_stats.xlsx'
		calculate_stats(gene_file, codon_file_list, tt, stats_output_file)
		combined_data_list = []
		for file in filtered_file_list:
			df = pd.read_excel(file)
			for index, row in df.iterrows():
				combined_data_list.append(row)
		combined_data = pd.DataFrame(combined_data_list, columns=df.columns)
		combined_file_name = f'processed_{species_name}_all_pep99.xlsx'
		combined_data.to_excel(combined_file_name, index=False)
		wb = openpyxl.load_workbook(combined_file_name)
		ws = wb.active

		custom_font = Font(name='Calibri', size=11, bold=False)
		no_border = Border(left=Side(style=None),
						right=Side(style=None),
						top=Side(style=None),
						bottom=Side(style=None))
		alignment = Alignment(horizontal='left', vertical='center')
		for cell in ws[1]:
			cell.font = custom_font
			cell.border = no_border
			cell.alignment = alignment

		wb.save(combined_file_name)
		xle_corr = None
		out_items = run_error_analysis(mutant_fasta_memory, combined_file_name, protein_dic_mem, gene_dic_mem, xle_seqs_mem, xle_corr, nsp_dic, wild_seqs_memory, mutant_seqs_memory, different_dna_set)
		nsp_dic = out_items[5]
		different_dna_set = out_items[8]
		if len(different_dna_set) > 0:
			with open(nsp_homology_file, 'wb') as f:
				pickle.dump(different_dna_set, f)
		workbook_out = out_items[1]
		format_matrix(out_items[3], remove_artifacts=True)
		codon_matrix(workbook_out, gene_file, protein_file, tt)
		codon_bias(workbook_out, gene_file, protein_file, tt, usage, mutant_fasta_memory)

		print("Finished")
		end_time = time.time()
		end_time_all = time.time()
		total_time = end_time - start_time
		total_time_all = end_time_all - start_time_all
		if total_time < 60:
			print(f"Total time for the combined dataset: {total_time:.2f} seconds")
		else:
			minutes = int(total_time // 60)
			seconds = int(total_time % 60)
			print(f"Total time for the combined dataset: {minutes} minute(s) {seconds} second(s)")

		if total_time_all < 60:
			print(f"Total time for all {total_files+1} files: {total_time_all:.2f} seconds")
		else:
			minutes = int(total_time_all // 60)
			seconds = int(total_time_all % 60)
			print(f"Total time for all {total_files+1} files : {minutes} minute(s) {seconds} second(s)")			