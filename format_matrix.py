import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
import sys
import getopt

#test for updates
def format_matrix(input_file, remove_artifacts=False):
	wb = openpyxl.load_workbook(input_file)
	ws = wb.active

	# Define number format and ranges
	decimal_format = '0.0000'
	cell_ranges_for_decimals = [ws['B24:U42'], ws['B48:V48']]  # List of ranges to format with decimals
	bold_center_range = ws['A23:V51']  # Range to apply bold and centered alignment
	regular_ranges = [ws['A45:V47'], ws['A50:B51'], ws['B23:U23'], ws['A24:A48'],  ws['V45:V48']]
	outline_ranges = [ws['B23:U23'], ws['A24:U42'], ws['B45:V45'], ws['A46:V48']]
	white_fill_ranges = [ws['A22:A23'], ws['B22:W22'], ws['V23:W44'], ws['W45:W49'], ws['A43:A45'], ws['B43:U44'], ws['A49:V49']]

	# Define font and alignment
	bold_font = Font(name='Calibri', bold=True)  # Set font to Calibri and bold
	regular_font = Font(name='Calibri', bold=False)
	center_alignment = Alignment(horizontal='center', vertical='center')

	# Bold and center
	for row in bold_center_range:
		for cell in row:
			cell.font = bold_font
			cell.alignment = center_alignment

	#Decimal formatting and Calibri font to the two matrices
	for cell_range in cell_ranges_for_decimals:
		for row in cell_range:
			for cell in row:
				cell.number_format = decimal_format
				cell.font = bold_font

	#3-color scale for conditional formatting
	color_scale_rule = ColorScaleRule(
		start_type='min', start_color='4BACC6',  # Blue for minimum
		mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Yellow for midpoint
		end_type='max', end_color='C0504D'  # Red for maximum
	)

	# Apply the color scale rule to each range
	ws.conditional_formatting.add('B24:U42', color_scale_rule)
	ws.conditional_formatting.add('B48:U48', color_scale_rule)

	# Remove artifacts
	grey_fill = PatternFill(fill_type='solid', start_color='D9D9D9', end_color='D9D9D9')
	if remove_artifacts:
		# Definitions can be changed
		ptm_definitions = {'NtoD', 'QtoE', 'EtoS', 'StoD', 'TtoE', 'StoA', 'YtoF'}
		for ptm in ptm_definitions:
			source = ptm[0]
			dest = ptm[-1]
			for row in ws['B23:U23']:
				for cell in row:
					if cell.value == source:
						source_loc = cell.column
						for row2 in ws['A24:A42']:
							for cell2 in row2:
								if cell2.value == dest:
									dest_loc = cell2.row
									target_cell = ws.cell(row=dest_loc, column=source_loc)
									target_cell.value = ''
									target_cell.fill = grey_fill
									

	for range in regular_ranges:
		regular_range = range
		for row in regular_range:
			for cell in row:
				cell.font = regular_font

	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)                

	for range in outline_ranges:
		outline_range = range
		for row in outline_range:
			for cell in row:
				cell.border = thin_border

	white_fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
	for range in white_fill_ranges:
		white_fill_range = range
		for row in white_fill_range:
			for cell in row:
				cell.fill = white_fill

	matrix_out = f"{input_file}"
	wb.save(matrix_out)
	return matrix_out

if __name__ == '__main__':
	input_file = ''
	remove_artifacts = False
	try:
		options, remainder = getopt.getopt(sys.argv[1:], '', ['input_file=', 'remove_artifacts'])
	except getopt.GetoptError:
		print("Usage: python format_matrix.py --input_file matrix.xlsx [--remove_artifacts]")
		sys.exit(2)

	for opt, arg in options:
		if opt == '--input_file':
			input_file = arg
		elif opt == '--remove_artifacts':
			remove_artifacts = True
		else:
			print(f"Warning! Command-line argument: {opt} not recognized. Exiting...")
			sys.exit(2)

	if not input_file:
		print('Error: --input_file argument is required.')
		sys.exit(2)

	format_matrix(input_file, remove_artifacts)
