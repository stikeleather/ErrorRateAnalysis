import sys
import getopt
import openpyxl
import re
from Bio import SeqIO
from Bio.Data import CodonTable
from generate_custom_database import fasta_to_dict

def find_mismatch(wt_seq, mut_seq):
	mismatch_pos = -1
	match_pos = -1

	# Loop over each possible starting position in wt_seq where mut_seq can fit
	for i in range(len(wt_seq) - len(mut_seq) + 1):
		segment = wt_seq[i:i+len(mut_seq)]
		mismatch_indices = [j for j in range(len(mut_seq)) if segment[j] != mut_seq[j]]
		
		# Check for one mismatch
		if len(mismatch_indices) == 1:
			mismatch_pos = i + mismatch_indices[0]  # position of mismatch in wt_seq
			match_pos = i  # starting position of alignment in wt_seq
			return (match_pos + 1), (mismatch_pos + 1), (match_pos + len(mut_seq))

	return -1, -1


def find_match(long_seq, gene_seq):
	try:
		start_index = long_seq.index(gene_seq)
		end_index = start_index + len(gene_seq)
		return start_index, end_index
	except ValueError:
		return -1, -1


def calculate_min_mismatches(target_string, string_list):
	min_mismatches = []
	min_positions = []
	for string in string_list:
		mismatch_count = sum(a != b for a, b in zip(target_string, string))
		min_mismatches.append(mismatch_count)

	minimum_mismatch = min(min_mismatches)
	for string in string_list:
		mismatch_count = sum(a != b for a, b in zip(target_string, string))
		if mismatch_count == minimum_mismatch:
			positions = [i + 1 for i, (a, b) in enumerate(zip(target_string, string)) if a != b]
			min_positions.append(positions)
	return minimum_mismatch, min_positions

def read_file_names(file_input):
	if isinstance(file_input, list):
		return file_input
	else:
		with open(file_input, 'r') as file:
			file_names = file.read().splitlines()
		return file_names


def codon_bias(input_file ,gene_file, protein_file, tt, usage):

	wild_table = {
	'TTT': 0, 'TTC': 0, 'TTA': 0, 'TTG': 0, 'TAT': 0, 'TAC': 0, 'TAA': 0, 'TAG': 0, 
	'CTT': 0, 'CTC': 0, 'CTA': 0, 'CTG': 0, 'CAT': 0, 'CAC': 0, 'CAA': 0, 'CAG': 0, 
	'ATT': 0, 'ATC': 0, 'ATA': 0, 'ATG': 0, 'AAT': 0, 'AAC': 0, 'AAA': 0, 'AAG': 0, 
	'GTT': 0, 'GTC': 0, 'GTA': 0, 'GTG': 0, 'GAT': 0, 'GAC': 0, 'GAA': 0, 'GAG': 0, 
	'TCT': 0, 'TCC': 0, 'TCA': 0, 'TCG': 0, 'TGT': 0, 'TGC': 0, 'TGA': 0, 'TGG': 0, 
	'CCT': 0, 'CCC': 0, 'CCA': 0, 'CCG': 0, 'CGT': 0, 'CGC': 0, 'CGA': 0, 'CGG': 0, 
	'ACT': 0, 'ACC': 0, 'ACA': 0, 'ACG': 0, 'AGT': 0, 'AGC': 0, 'AGA': 0, 'AGG': 0, 
	'GCT': 0, 'GCC': 0, 'GCA': 0, 'GCG': 0, 'GGT': 0, 'GGC': 0, 'GGA': 0, 'GGG': 0}

	error_table = {
	'TTT': 0, 'TTC': 0, 'TTA': 0, 'TTG': 0, 'TAT': 0, 'TAC': 0, 'TAA': 0, 'TAG': 0, 
	'CTT': 0, 'CTC': 0, 'CTA': 0, 'CTG': 0, 'CAT': 0, 'CAC': 0, 'CAA': 0, 'CAG': 0, 
	'ATT': 0, 'ATC': 0, 'ATA': 0, 'ATG': 0, 'AAT': 0, 'AAC': 0, 'AAA': 0, 'AAG': 0, 
	'GTT': 0, 'GTC': 0, 'GTA': 0, 'GTG': 0, 'GAT': 0, 'GAC': 0, 'GAA': 0, 'GAG': 0, 
	'TCT': 0, 'TCC': 0, 'TCA': 0, 'TCG': 0, 'TGT': 0, 'TGC': 0, 'TGA': 0, 'TGG': 0, 
	'CCT': 0, 'CCC': 0, 'CCA': 0, 'CCG': 0, 'CGT': 0, 'CGC': 0, 'CGA': 0, 'CGG': 0, 
	'ACT': 0, 'ACC': 0, 'ACA': 0, 'ACG': 0, 'AGT': 0, 'AGC': 0, 'AGA': 0, 'AGG': 0, 
	'GCT': 0, 'GCC': 0, 'GCA': 0, 'GCG': 0, 'GGT': 0, 'GGC': 0, 'GGA': 0, 'GGG': 0}

	wb = openpyxl.load_workbook(input_file)
	ws = wb["Sheet1"]
	
	aa_codons = CodonTable.unambiguous_dna_by_id[tt]
	stop_codons = aa_codons.stop_codons
	aa_codons = dict(aa_codons.forward_table.items())
	for stop in stop_codons:
		aa_codons[stop] = ''

	#Keys are amino acids and values are codons in a list
	distance_lib = dict()
	for k, v in aa_codons.items():
		if v not in distance_lib:
			distance_lib[v] = [k]
		else:
			distance_lib[v].append(k)

	codon_table = {
		'TTT': 0, 'TTC': 0, 'TTA': 0, 'TTG': 0, 'TAT': 0, 'TAC': 0, 'TAA': 0, 'TAG': 0, 
		'CTT': 0, 'CTC': 0, 'CTA': 0, 'CTG': 0, 'CAT': 0, 'CAC': 0, 'CAA': 0, 'CAG': 0, 
		'ATT': 0, 'ATC': 0, 'ATA': 0, 'ATG': 0, 'AAT': 0, 'AAC': 0, 'AAA': 0, 'AAG': 0, 
		'GTT': 0, 'GTC': 0, 'GTA': 0, 'GTG': 0, 'GAT': 0, 'GAC': 0, 'GAA': 0, 'GAG': 0, 
		'TCT': 0, 'TCC': 0, 'TCA': 0, 'TCG': 0, 'TGT': 0, 'TGC': 0, 'TGA': 0, 'TGG': 0, 
		'CCT': 0, 'CCC': 0, 'CCA': 0, 'CCG': 0, 'CGT': 0, 'CGC': 0, 'CGA': 0, 'CGG': 0, 
		'ACT': 0, 'ACC': 0, 'ACA': 0, 'ACG': 0, 'AGT': 0, 'AGC': 0, 'AGA': 0, 'AGG': 0, 
		'GCT': 0, 'GCC': 0, 'GCA': 0, 'GCG': 0, 'GGT': 0, 'GGC': 0, 'GGA': 0, 'GGG': 0}
	
	sequence_dic = fasta_to_dict(gene_file)
	# Calculate codon frequencies
	for gene in sequence_dic:
		sequence = sequence_dic[gene]
		for nt in range(0, len(sequence), 3):
			codon = sequence[nt:nt+3]
			codon_table[codon] += 1

	total_codons = sum(codon_table.values())

	if usage == 'abs':
		for codon in codon_table:
			codon_table[codon] = round(codon_table[codon] / total_codons, 4)
	elif usage == 'rel':
		for aa in distance_lib:
			family_codon_total = 0
			for codon in distance_lib[aa]:
				family_codon_total += codon_table[codon]
			for codon in distance_lib[aa]:
				codon_table[codon] = round(codon_table[codon] / family_codon_total, 2)

	row_num = 1
	for row in range(2,ws.max_row+1):
		if ws.cell(row,22).value is not None:
			row_num += 1
		else:
			break
	row_num = row_num+1

	
	dash_pattern = re.compile(r'[-]')
	dash_dot_pattern = re.compile(r'[-.]')
	digit_pattern = re.compile(r'\d+')

	#Create output workbook
	wb1 = openpyxl.Workbook()
	wb1.create_sheet(index=0, title='Sheet1')
	ws1 = wb1["Sheet1"]
	ws1.cell(1,1).value = 'accession'
	ws1.cell(1,2).value = 'sequence'
	ws1.cell(1,3).value = 'codon_usage'
	ws1.cell(1,4).value = 'psm_count'
	ws1.cell(1,5).value = 'min_distance'
	ws1.cell(1,6).value = 'AA'
	ws1.cell(1,7).value = 'codon'
	ws1.cell(1,8).value = 'usage'
	ws1.cell(1,9).value = 'error_rate'


	

	ws1.cell(1,10).value = 'mut1pos1'
	ws1.cell(1,11).value = 'mut1pos2'
	ws1.cell(1,12).value = 'mut1pos3'
	ws1.cell(1,13).value = 'mut2pos12'
	ws1.cell(1,14).value = 'mut2pos23'
	ws1.cell(1,15).value = 'mut2pos13'
	ws1.cell(1,16).value = 'mut3pos123'
	ws1.cell(1,17).value = 'min_dis1'
	ws1.cell(1,18).value = 'min_dis2'
	ws1.cell(1,19).value = 'min_dis3'

	#########################################################################################
	protein_dic = fasta_to_dict(protein_file)

	row_num = 1
	for row in range(2,ws.max_row+1):
		if ws.cell(row,1).value is not None:
			row_num += 1
		else:
			break
	row_num = row_num+1


	row_pos = 2
	mut1_pos1 = 0
	mut1_pos2 = 0
	mut1_pos3 = 0
	mut2_pos12 = 0
	mut2_pos23 = 0
	mut2_pos13 = 0
	mut3 = 0
	min1 = 0
	min2 = 0
	min3 = 0
	for row in range(2,row_num):
		if 'wild' not in ws.cell(row,1).value:
			gene = ws.cell(row,1).value
			split_gene = re.split(dash_dot_pattern,gene)
			mutation = split_gene[4]
			mutation_pos = int(re.findall(digit_pattern, mutation)[0])
			mutation = mutation[-1]
			gene_id = split_gene[1]
			# print(gene_id)
			nt_sequence = sequence_dic[gene_id]
			protein_seq = protein_dic[gene_id]
			aa_sequence = ws.cell(row,2).value
			mismatches = find_mismatch(protein_seq, aa_sequence)
			nt_codon = ''
			nt_pos = mismatches[1] * 3
			# print(nt_pos)
			# print(len(nt_sequence))
			if nt_pos >= 0:
				for i in range(3):
					nt_codon += nt_sequence[nt_pos-3+i]
				min_mismatches = calculate_min_mismatches(nt_codon,distance_lib[mutation])
				ws1.cell(row_pos,1).value = gene
				ws1.cell(row_pos,2).value = aa_sequence
				ws1.cell(row_pos,3).value = codon_table[nt_codon]
				psm_count = ws.cell(row,3).value
				ws1.cell(row_pos,4).value = psm_count
				ws1.cell(row_pos,5).value = min_mismatches[0]
				error_table[nt_codon] += psm_count
				codon_count = 0
				# print(gene)
				# print(aa_sequence)
				# print(nt_codon)
				# print(nt_pos)
				# print(mismatches)
				# print(min_mismatches)

				# Count Position Data
				pos_data = min_mismatches[1]
				if min_mismatches[0] == 1:
					min1 += psm_count
					unique_pos1 = set(tuple(sublist) for sublist in pos_data)
					for item in unique_pos1:
						if item == (1,):
							mut1_pos1 += psm_count
						elif item == (2,):
							mut1_pos2 += psm_count
						elif item == (3,):
							mut1_pos3 += psm_count
				elif min_mismatches[0] == 2:
					min2 += psm_count
					unique_pos2 = set(tuple(sublist) for sublist in pos_data)
					for item in unique_pos2:
						if item == (1, 2):
							mut2_pos12 += psm_count
						elif item == (2, 3):
							mut2_pos23 += psm_count
						elif item == (1, 3):
							mut2_pos13 += psm_count
				elif min_mismatches[0] == 3:
					min3 += psm_count
					mut3 += psm_count
				# if (mismatches[2]*3) == len(nt_sequence):
				# 	print("gene_length")
				# 	print(len(nt_sequence))
				# 	print("mismatches[2]*3")
				# 	print((mismatches[2]*3))	
				for nt in range((mismatches[0]*3)-3, (mismatches[2]*3), 3):
					codon_count += 1
					codon = nt_sequence[nt:nt+3]
					# print(codon)
					if codon_count != mutation_pos:
						# print(mutation_pos)
						# print(codon)
						wild_table[codon] += psm_count
				row_pos += 1

		else:
			gene = ws.cell(row,1).value
			try:
				#print(gene)
				split_gene = re.split(dash_pattern,gene)
				gene_id = split_gene[1]
				aa_sequence = ws.cell(row,2).value
				#print(f"{aa_sequence}\n")
				nt_sequence = sequence_dic[gene_id]
				#print(f"{nt_sequence}\n")
				#print(len(nt_sequence))
				protein_seq = protein_dic[gene_id]
				#print(f"{protein_seq}\n")
				#print(len(protein_seq))
				matches = find_match(protein_seq, aa_sequence)
				# if (matches[1]*3) == len(nt_sequence):
				# 	print("gene_length")
				# 	print(len(nt_sequence))
				# 	print("matches[1]*3")
				# 	print(matches[1]*3)
				#print(matches[0])
				#print("matches[1]*3")
				#print(matches[1]*3)
				psm_count = ws.cell(row,3).value
				for nt in range((matches[0]*3),(matches[1]*3),3):
					codon = nt_sequence[nt:nt+3]
					#print(codon)
					wild_table[codon] += psm_count
			except KeyError:
				print(f"Error in codon bias: {gene}, check the protein and gene sequence, then the genbank file entry")
			
	row_pos = 2
	for key in wild_table:
		if aa_codons[key] != '':
			ws1.cell(row_pos,6).value = aa_codons[key]
			ws1.cell(row_pos,7).value = key
			ws1.cell(row_pos,8).value = codon_table[key]
			codon_wilds = wild_table[key]
			codon_errors = error_table[key]
			codon_total = codon_wilds + codon_errors
			if codon_total != 0:
				ws1.cell(row_pos,9).value = codon_errors/codon_total
			else:
				ws1.cell(row_pos,9).value = 0
			row_pos += 1
	
	ws1.cell(2,10).value = mut1_pos1
	ws1.cell(2,11).value = mut1_pos2
	ws1.cell(2,12).value = mut1_pos3
	ws1.cell(2,13).value = mut2_pos12
	ws1.cell(2,14).value = mut2_pos23
	ws1.cell(2,15).value = mut2_pos13
	ws1.cell(2,16).value = mut3
	ws1.cell(2,17).value = min1
	ws1.cell(2,18).value = min2
	ws1.cell(2,19).value = min3

	for row in range(2,63):
		codon = ws1.cell(row,7).value
		codon = codon.replace('T', 'U')
		ws1.cell(row,7).value = codon

	ws1.insert_cols(8)
	ws1.cell(1,8).value = 'label'
	for row in range(2,63):
		amino = ws1.cell(row,6).value
		codon = ws1.cell(row,7).value
		ws1.cell(row,8).value = f"{amino}, {codon}"

	ws1.insert_cols(11)
	ws1.cell(1,11).value = 'codon_total'
	row_pos = 2
	for key in wild_table:
		if aa_codons[key] != '':
			codon_wilds = wild_table[key]
			codon_errors = error_table[key]
			codon_total = codon_wilds + codon_errors
			ws1.cell(row_pos,11).value = codon_total
			row_pos += 1

	ws1.cell(1,22).value = 'translation_error'
	# ws1.cell(2,22).value = ws.cell(2,30).value
	#for update
	ws1.cell(2,22).value = ws.cell(2,8).value


	#wb1.save("codon_bias_global_"+input_file)

	#Use if comparing global vs relative usage
	if usage == 'abs':
		output_file = f"codon_bias_absolute_{input_file}"
		wb1.save(output_file)
	elif usage == 'rel':
		output_file = f"codon_bias_relative_{input_file}"
		wb1.save(output_file)

	return output_file

def print_usage():
	print('Usage:')
	print('Example 1: python codon_bias.py --input_file analyzed_processed_coli_pep99.xlsx --gene_file genes.fasta --protein_file proteome.fasta --tt <1-33> --usage abs')
	print('Example 2: python codon_bias.py --file_list file_list.txt --gene_file genes.fasta --protein_file proteome.fasta --tt <1-33> --usage abs')
	print('Arguments:')
	print('--input_file    : The processed output file from run_error_analysis')
	print('--file_list     : Text (.txt) file containing a list of input files')
	print('--gene_file     : Gene sequences in FASTA format')
	print('--protein_file  : Protein sequences in FASTA format')
	print('--tt            : Translation table (e.g., 11)')
	print('--usage         : Codon usage type, "abs" for absolute or "rel" for relative usage')
	sys.exit(0)

if __name__ == '__main__':
	file_list = None
	try:
		options, remainder = getopt.getopt(sys.argv[1:], 'h', ['input_file=', 'file_list=', 'gene_file=', 'protein_file=', 'tt=', 'usage=', 'help'])
	except getopt.GetoptError:
		print_usage()

	input_file = None
	gene_file = None
	protein_file = None
	usage = None
	tt = None

	for opt, arg in options:
		if opt in ('-h', '--help'):
			print_usage()
		elif opt == '--input_file':
			input_file = arg
		elif opt == '--file_list':
			file_list = arg
		elif opt == '--gene_file':
			gene_file = arg
		elif opt == '--protein_file':
			protein_file = arg
		elif opt == '--tt':
			tt = int(arg)
		elif opt == '--usage':
			usage = arg

	if usage not in ['abs', 'rel']:
		print("Error: The 'usage' argument must be either 'abs' or 'rel'.")
		sys.exit(2)

	if file_list:
		input_files = read_file_names(file_list)
	else:
		input_files = [input_file]

	if not input_file and not file_list:
		print("Error: Missing input_file or file_list argument")
		sys.exit(2)
	elif not gene_file:
		print("Error: Missing gene_file argument")
		sys.exit(2)
	elif not protein_file:
		print("Error: Missing protein_file argument")
		sys.exit(2)
	elif not tt:
		print("Error: Missing translation table (tt) argument")
		sys.exit(2)

	for input_file in input_files:
		codon_bias(input_file, gene_file, protein_file, tt, usage)
